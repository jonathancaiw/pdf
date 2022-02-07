# pip install pdfminer.six
import os
from io import StringIO
from openpyxl import load_workbook
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
from util import *

CONTRACT_CONFIG = '/Users/caiwei/Desktop/contract_config.xlsx'  # 合同配置文件
FILE_PATH = '/Users/caiwei/Desktop/2022-01-01后合同已签数据-生产环境'  # 合同路径
KEY_VALUES = {'contract_start_date': '合同开始日期', 'contract_end_date': '合同结束日期', 'charge_end_date': '收费结束日期'}  # 检查项

global total
global mismatch
global process
global resume
total = 0  # 合同计数
mismatch = 0  # 日期不匹配计数
process = True
resume = '861615'


def get_config(filename):
    """
    加载合同配置文件
    :param filename:
    :return:
    """
    config = {}

    workbook = load_workbook(filename)
    worksheet = workbook[workbook.sheetnames[0]]

    for row in worksheet.rows:
        if row[2].value is None:
            write_log(row[1].value + 'contract_start_date is None')
            continue

        if row[3].value is None:
            write_log(row[1].value + 'contract_end_date is None')
            continue

        config[row[0].value] = {'organization_name': row[1].value,
                                'contract_start_date': row[2].value[0:10],
                                'contract_end_date': row[3].value[0:10],
                                'charge_end_date': None if row[4].value is None else row[4].value[0:10]}

    workbook.close()

    return config


def format_date(date):
    """
    日期格式化
    :param date:
    :return:
    """
    if date is None:
        return None
    else:
        return date[:4] + '年' + str(int(date[5:7])) + '月' + str(int(date[8:])) + '日'


def parse_pdf(filename):
    """
    处理pdf
    :param filename:
    :return:
    """
    global resume
    global process

    org_no = filename.split('/')[-2]  # 组织机构编码

    if not process:
        if org_no != resume:
            return
        else:
            process = True

    if org_no not in contract_config.keys():
        write_log(org_no + ' contract config is None')
        return

    with open(filename, 'rb') as in_file:
        parser = PDFParser(in_file)
        doc = PDFDocument(parser)
        rsrcmgr = PDFResourceManager()
        device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        for page in PDFPage.create_pages(doc):
            interpreter.process_page(page)

    text = output_string.getvalue()

    check_list = {}
    for keyword in KEY_VALUES.keys():
        value = format_date(contract_config[org_no][keyword])
        if value is not None and len(value.strip()) > 0:
            if value in check_list.keys():
                check_list[value].append(keyword)
            else:
                check_list[value] = [keyword]

    check_contract(filename, text, check_list)


def check_contract(filename, text, check_list):
    """
    检查
    :param filename:
    :param text:
    :param check_list:
    :return:
    """
    global mismatch

    for value, keywords in check_list.items():
        index = -1
        for keyword in keywords:
            index = text.find(value, 0 if index == -1 else index + 1)
            if index == -1:
                write_log('%s,%s,%s' % (KEY_VALUES[keyword], value, filename))
                mismatch += 1
                return


def find_pdf(file_path):
    """
    递归查找pdf
    :param file_path:
    :return:
    """
    global total
    for filename in os.listdir(file_path):
        sub_path = file_path + '/' + filename
        if os.path.isdir(sub_path):
            find_pdf(sub_path)
        elif filename.endswith('.pdf'):
            parse_pdf(sub_path)
            print(filename)
            total += 1
        else:
            print('other file', sub_path)

        print('match %d, mismatch %d, total %d ...' % (total - mismatch, mismatch, total))


output_string = StringIO()

"""
SELECT
	organization_no,
	organization_name,
	contract_start_date,
	contract_end_date,
	charge_end_date 
FROM
	"land_contract"."lc_farm_contract_config" 
WHERE
	"status_cd" = '1' 
	AND "year_no" = '2022'
"""
contract_config = get_config(CONTRACT_CONFIG)

find_pdf(FILE_PATH)

write_log('match %d, mismatch %d, total %d' % (total - mismatch, mismatch, total))

from openpyxl import load_workbook
from util import *

PDF_CHECK_XLSX = '/Users/caiwei/Desktop/log_20220207_pdf_check.xlsx'
CONTRACT_XLSX = '/Users/caiwei/Desktop/2022-01-01后合同已签数据-生产环境/2022-01-01后合同已签数据.xlsx'


def read_pdf_check(filename):
    """
    加载合同检查文件
    :param filename:
    :return:
    """
    pdf_check = {}  # 处理日期、问题类型、正确值、合同路径

    workbook = load_workbook(filename)
    worksheet = workbook[workbook.sheetnames[0]]

    skip_header = True

    for row in worksheet.rows:
        if skip_header:
            skip_header = False
            continue

        file_path = row[5].value
        splits = file_path.split('.')[0].split('-')
        key = splits[-1]
        pdf_check[key] = {'date': row[0].value,
                          'type': row[3].value,
                          'label': row[4].value,
                          'file_path': row[5].value}

    workbook.close()

    return pdf_check


def read_contract(filename):
    """
     加载合同文件
     :param filename:
     :return:
     """
    contract = {}

    workbook = load_workbook(filename)
    worksheet = workbook[workbook.sheetnames[0]]

    header = {}

    for row in worksheet.rows:
        if len(header) == 0:
            for i in range(0, len(row)):
                header[row[i].value] = i
            continue

        contract[row[header['fdd_contract_no']].value] = {'contract_sign_no': row[header['contract_sign_no']].value,
                                                          'farm_sign_no': row[header['farm_sign_no']].value,
                                                          'organization_no': row[header['organization_no']].value,
                                                          'organization_name': row[header['organization_name']].value,
                                                          'name': row[header['farmer_name']].value,
                                                          'id_number': row[header['farmer_id_number']].value}

    workbook.close()

    return contract


pdf_check = read_pdf_check(PDF_CHECK_XLSX)

contract = read_contract(CONTRACT_XLSX)

for key in pdf_check.keys():
    if key in contract.keys():
        write_log('%s,%s,%s,%s,%s,%s,%s,%s,%s,%s' % (
            key, pdf_check[key]['type'], pdf_check[key]['label'], contract[key]['organization_no'],
            contract[key]['organization_name'], contract[key]['name'], contract[key]['id_number'],
            '种植户已签字' if contract[key]['contract_sign_no'] == 1 else '种植户未签字',
            '农场已盖章' if contract[key]['farm_sign_no'] == 2 else '农场未盖章', pdf_check[key]['file_path']))
    else:
        write_log(pdf_check[key]['file_path'] + ' doesn''t exists')
